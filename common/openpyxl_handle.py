import time

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

"""
1.打开表单
2.读取表头
3.读取所有数据 （list嵌套dict的可读性强）
4.指定单元格写入数据（使用静态方法，不要用实例方法）
"""


class ExcelHandler:
    def __init__(self, file):
        """初始化函数"""
        self.file = file
        # 这里也可以把name放到init中，但是一次只能操作1个表单，先不用。
        # 0.35 28分钟前
        # 0.36 20分钟，self.wb = openpyxl.load_workbook(file),不推荐这么写，因为每次打开原来这个，下面函数修改会不生效。

    def open_sheet(self, name) -> Worksheet:
        """打开表单
        在函数后加-> 代表函数的返回类型 ，是一种函数注解
        作用：小技巧，解决 sheet = wb[name] 后，在pycharm中，无法sheet点操作
        """

        '''
        # 这里视频没有展示write1的实例方法（只有静态方法write2）
        #!!!如下：wb和sheet不加self，会提示AttributeError: 'ExcelHandler' object has no attribute 'wb'
        wb = openpyxl.load_workbook(self.file)
        sheet = wb[name]
        wb.close()  # 一般读文件会自动关闭，写文件不关闭会造成数据混乱
        return sheet
        '''

        # 修改加入self后如下，且不影响get_*方法
        self.wb = openpyxl.load_workbook(self.file)
        self.sheet = self.wb[name]
        self.wb.close()  # 一般读文件会自动关闭，写文件不关闭会造成数据混乱
        return self.sheet

    def get_title(self, name):
        """获取首行，返回首行的list形式"""
        sheet = self.open_sheet(name)
        title = []
        for cell in sheet[1]:
            title.append(cell.value)  # 这里的value在点的时候没有提示，见openpyxl_basic.py的说明
        return title

    def get_all(self, name):
        """获取所有数据"""
        sheet = self.open_sheet(name)
        rows = list(sheet.rows)[1:]  # 第1行是表头,对应的是list的第0行

        # 数据类型1：把数据放在list套list中。但可读性不高
        # data = []
        # for column in rows:
        #     row_data = []
        #     for cell in column:
        #         row_data.append(cell.value)
        #     data.append(row_data)
        # return data

        # # 数据类型2：把数据放在list套dict中。方便查找
        data = []
        title = self.get_title(name)
        for row in rows:
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            # print(row_data)  # 每次输出excel的一行数据

            data_dict = dict(zip(title, row_data))  # zip这行数据
            data.append(data_dict)  # 把每次的dict放到data中
        return data

    def write1(self, sheet_name, row, col, value):
        """
        写入excel数据
        """
        sheet = self.open_sheet(sheet_name)
        sheet.cell(row, col).value = value
        self.wb.save(self.file)
        self.wb.close()

    @staticmethod
    def write2(file, sheet_name, row, col, data):
        """
        写入excel数据
        """
        wb = openpyxl.load_workbook(file)
        sheet = wb[sheet_name]
        sheet.cell(row, col).value = data
        wb.save(file)  # 必须save，否则数据没保存
        wb.close()


# 用来先测试封装的内容：
if __name__ == '__main__':
    localtime = time.asctime(time.localtime(time.time()))

    excel = ExcelHandler(r"excel.xlsx")  # r"excel.xlsx"
    excel.write1("Sheet1", 4, 3, localtime)
    # excel.write2(r"excel.xlsx", "Sheet1", 4, 3, localtime)
    print(excel.get_all("Sheet1"))
